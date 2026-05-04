import openpyxl
from pathlib import Path

excel_path = Path(r"c:\Users\User\Downloads\test_automation\test_automation\IT23401044_Assignment_1_Test_cases_50_FAIL.xlsx")
if not excel_path.exists():
    print(f"File not found: {excel_path}")
else:
    wb = openpyxl.load_workbook(excel_path)
    print(f"Sheet names: {wb.sheetnames}")
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
